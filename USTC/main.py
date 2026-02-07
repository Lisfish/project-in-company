"""
M50成品锁报告生成工具
作者：严江阳
日期：2024年
"""

import os
import re
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import config  # 导入配置文件
import sys


class InspectionReportGenerator:
    def __init__(self):
        self.wb = None
        self.template_path = None
        self.images_data = []
        self.defect_images = []

        # 抽样计划数据
        self.sampling_plan = {
            'ranges': [(151, 280), (281, 500), (501, 1200), (1201, 3200),
                       (3201, 10000), (10001, 35000), (35001, float('inf'))],
            'sample_sizes': [13, 20, 32, 50, 80, 125, 200],
            'critical': [0, 0, 0, 0, 0, 0, 0],
            'major': [0, 0, 0, 1, 2, 3, 5],
            'minor': [0, 1, 2, 3, 5, 7, 10]
        }

    def load_template(self, template_path):
        """加载Excel模板"""
        try:
            self.wb = openpyxl.load_workbook(template_path)
            self.template_path = template_path
            print(f"✓ 模板加载成功: {Path(template_path).name}")
            return True
        except Exception as e:
            print(f"✗ 加载模板失败: {e}")
            return False

    def fill_basic_info(self, data):
        """
        填充基本信息
        data格式: {
            'inspector': '张三',
            'inspection_date': '2024/03/15',
            'po_number': 'PO-2024-1234',
            'sku': 'P61718/M50XTCCSEN',
            'ship_date': '2024/03/20',
            'ship_quantity': 1800,
            'report_no': 'OI2024-001',
            'customer': 'Master Lock',
            'drawing_no': '64678 Rev.J',
            'approver': 'Gary Tu',
            'approval_date': '2024/03/15'
        }
        """
        try:
            ws = self.wb['出货检查表']

            # 填充基本信息
            if 'inspector' in data:
                ws['C4'] = data['inspector']  # 检验员
            if 'inspection_date' in data:
                ws['G4'] = data['inspection_date']  # 检验日期
            if 'po_number' in data:
                ws['G5'] = data['po_number']  # 客户订单号
                ws['B53'] = f'See tab" Reference pictures {data["po_number"]}"'
            if 'sku' in data:
                ws['C6'] = data['sku']  # 料号
                ws['C17'] = f"BOM {data['sku'] } Rev E  ECO-017206"
            if 'ship_date' in data:
                ws['B7'] = f"计划出货日期：{data['ship_date']}"# 计划出货日期
            if 'ship_quantity' in data:
                ws['G7'] = data['ship_quantity']  # 出货数量
            if 'report_no' in data:
                ws['B3'] = f'出货检查报告编号 {data["report_no"]}'  # 报告编号
            if 'customer' in data:
                ws['C5'] = data['customer']
                # F5 填写客户图纸及版本号
            if 'sku' in data:
                sku = data['sku']
                drawing_no = ""
                for model, drawing in config.DRAWING_RULES.items():
                    if model in sku:
                        drawing_no = drawing
                        break
                if not drawing_no:
                    drawing_no = data.get('drawing_no', '')
                ws['F6'] = f"客户图纸及版本号：{drawing_no}"
                ws['C16'] = f"Master Lock drawing: {drawing_no}"

            if 'inspector' in data and 'inspection_date' in data:
                ws['D49'] = f"{data['inspector']}/{data['inspection_date']}"
                # C50 填写批准人信息
            if 'approver' in data:
                ws['C50'] = data['approver']
                # D50 填写批准人签名/日期（格式：批准人签名/日期：批准人/批准日期）
            if 'approver' in data and 'approval_date' in data:
                ws['D50'] = f"批准人签名/日期：{data['approver']}/{data['approval_date']}"
            # 更新抽样计划
            if 'ship_quantity' in data:
                self.update_sampling_plan(data['ship_quantity'])

            print("✓ 基本信息填充完成")
            return True

        except Exception as e:
            print(f"✗ 填充基本信息失败: {e}")
            return False

    def update_sampling_plan(self, quantity):
        try:
            ws = self.wb['出货检查表']

            range_col_mapping = {
                (151, 280): 'C',
                (281, 500): 'D',
                (501, 1200): 'E',
                (1201, 3200): 'F',
                (3201, 10000): 'G',
                (10001, 35000): 'H',
                (35001, float('inf')): 'I'
            }

            # 2. 定义行索引（按你的表格结构）
            row_mapping = {
                'lot_quantity': 10,  # Lot quantity行
                'sample_size': 11,  # Sample Size行
                'critical': 12,  # Critical [0]行
                'major': 13,  # Major [1.0]行
                'minor': 14  # Minor [2.5]行
            }

            # 3. 遍历预设区间，匹配数量并处理
            for i, (min_qty, max_qty) in enumerate(self.sampling_plan['ranges']):
                if min_qty <= quantity <= max_qty:
                    # 获取匹配的列
                    target_col = range_col_mapping[(min_qty, max_qty)]

                    # 4. 填写Lot quantity
                    lot_quantity_cell = f"{target_col}{row_mapping['lot_quantity']}"
                    ws[lot_quantity_cell] = quantity

                    # 5. 填写抽样数据到对应行
                    # Sample Size
                    sample_size_cell = f"{target_col}{row_mapping['sample_size']}"
                    ws[sample_size_cell] = self.sampling_plan['sample_sizes'][i]
                    # Critical [0]
                    critical_cell = f"{target_col}{row_mapping['critical']}"
                    ws[critical_cell] = self.sampling_plan['critical'][i]
                    # Major [1.0]
                    major_cell = f"{target_col}{row_mapping['major']}"
                    ws[major_cell] = self.sampling_plan['major'][i]
                    # Minor [2.5]
                    minor_cell = f"{target_col}{row_mapping['minor']}"
                    ws[minor_cell] = self.sampling_plan['minor'][i]

                    # 高亮标红
                    from openpyxl.styles import Font, PatternFill
                    red_font = Font(color="FF0000", size=8, bold=False)  # 红色字体
                    # 批量设置样式
                    for cell in [sample_size_cell, critical_cell, major_cell, minor_cell]:
                        ws[cell].font = red_font

                    print(
                        f"✓ 抽样计划更新: 数量={quantity}, 写入列={target_col}, 样本数={self.sampling_plan['sample_sizes'][i]}")
                    return True

            # 若数量不在预设区间（如≤150）
            print(f"⚠ 出货数量 {quantity} 不在有效范围内")
            return False

        except Exception as e:
            print(f"✗ 更新抽样计划失败: {e}")
            return False

    def add_defect_records(self, defects):
        """
        添加缺陷记录（先取消合并→写入数据→重新合并单元格）
        defects格式: [
            {'description': '划痕', 'critical': 0, 'major': 1, 'minor': 0},
            {'description': '颜色不均', 'critical': 0, 'major': 0, 'minor': 1}
        ]
        """
        try:
            ws = self.wb['出货检查表']

            # 缺陷记录起始行/结束行（最多8条）
            start_row = 21  # 第21行开始是缺陷记录
            end_row = start_row + 7  # 8条记录：21-28行

            # 取消合并单元格
            merge_ranges_to_restore = []
            for merge_range in list(ws.merged_cells.ranges):
                if not (merge_range.max_row < start_row or merge_range.min_row > end_row):
                    merge_ranges_to_restore.append({
                        'range': str(merge_range),  # 合并范围字符串（如"C21:F21"）
                        'min_col': merge_range.min_col,
                        'max_col': merge_range.max_col,
                        'min_row': merge_range.min_row,
                        'max_row': merge_range.max_row
                    })
                    # 取消合并
                    ws.unmerge_cells(str(merge_range))

            # 写入缺陷数据
            for i, defect in enumerate(defects):
                if i >= 8:  # 最多8条记录
                    break

                row = start_row + i
                ws[f'B{row}'] = i + 1  # 序号
                ws[f'C{row}'] = defect.get('description', '')  # 缺陷品描述
                ws[f'G{row}'] = defect.get('critical', 0)  # 致命缺陷数量
                ws[f'H{row}'] = defect.get('major', 0)  # 严重缺陷数量
                ws[f'I{row}'] = defect.get('minor', 0)  # 轻微缺陷数量

            # 重新合并单元格
            for merge_info in merge_ranges_to_restore:
                ws.merge_cells(merge_info['range'])

            print(f"✓ 添加了 {min(len(defects), 8)} 条缺陷记录")
            return True

        except Exception as e:
            print(f"✗ 添加缺陷记录失败: {e}")
            return False

    def scan_images_folder(self, folder_path):
        """扫描图片文件夹，按步骤分类"""
        try:
            self.images_data = []
            self.defect_images = []
            image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.gif']

            # 定义步骤关键词
            step_keywords = {
                'Step 1': ['step1', 'step 1', '步骤1', 'shipping', '外箱', 'case', '标签'],
                'Step 2': ['step2', 'step 2', '步骤2', 'edge', 'sharp', '毛边', '边缘'],
                'Step 3': ['step3', 'step 3', '步骤3', 'key code', 'oval', '钥匙标签', '标签'],
                'Step 4': ['step4', 'step 4', '步骤4', 'nut', 'push', '螺母', '固定'],
                'Step 5': ['step5', 'step 5', '步骤5', 'unlock', 'key', 'shackle', '锁具', '测试'],
                'Step 5（1）': ['step5_1', 'step5(1)', '步骤5(1)'],
                'Step 5（2）': ['step5_2', 'step5(2)', '步骤5(2)'],
                'Step 5（3）': ['step5_3', 'step5(3)', '步骤5(3)'],
                'Step 5（4）': ['step5_4', 'step5(4)', '步骤5(4)'],
                'Step 5（5）': ['step5_5', 'step5(5)', '步骤5(5)']
            }

            for file_path in Path(folder_path).glob('*'):
                if file_path.suffix.lower() in image_extensions:
                    filename = file_path.name.lower()

                    for defect_word in config.DEFECT_WORDS:
                        if defect_word.lower() in filename:
                            self.defect_images.append(str(file_path))
                            break
                    # 确定图片对应的步骤
                    assigned_step = None
                    for step, keywords in step_keywords.items():
                        if any(keyword in filename for keyword in keywords):
                            assigned_step = step
                            break

                    if not assigned_step:
                        # 尝试从文件名中提取步骤信息
                        step_match = re.search(r'step[_\s]*(\d+)', filename)
                        if step_match:
                            assigned_step = f"Step {step_match.group(1)}"
                        else:
                            assigned_step = "Step 1"  # 默认

                    self.images_data.append({
                        'path': str(file_path),
                        'filename': file_path.name,
                        'step': assigned_step
                    })

            # 按步骤排序
            self.images_data.sort(key=lambda x: (
                x['step'].replace('Step ', ''),
                x['step'].replace('（', '').replace('）', '')
            ))

            print(f"✓ 扫描到 {len(self.images_data)} 张图片")
            print(f"✓ 扫描到 {len(self.defect_images)} 张缺陷图片")
            return self.images_data

        except Exception as e:
            print(f"✗ 扫描图片文件夹失败: {e}")
            return []

    def _insert_defect_images(self):
        """将所有标记为缺陷的图片以 2xN 网格形式插入，横向跨度为 B-E 和 F-I"""
        # 1. 严格去重：使用 unique_defect_images 作为统一变量名
        unique_defect_images = list(dict.fromkeys(self.defect_images))

        if not unique_defect_images:
            return

        try:
            ws = self.wb.active
            cfg = config.DEFECT_IMAGE_CONFIG
            from openpyxl.utils import column_index_from_string
            from openpyxl.utils import get_column_letter

            # 获取起始位置
            start_col = column_index_from_string("B")  # B列
            start_row = 55

            for i, img_path in enumerate(unique_defect_images):
                if not os.path.exists(img_path):
                    continue

                # 计算网格位置 (i=0 是第一张, i=1 是第二张...)
                row_idx = i // 2  # 每行2张
                col_idx = i % 2  # 0代表左边，1代表右边

                # 计算具体的行列坐标
                # cfg["row_span"] 应该是 14 (54到67行)
                # cfg["col_span"] 应该是 4 (B到E是4列)
                current_row = start_row + (row_idx * cfg["row_span"])
                current_col = start_col + (col_idx * cfg["col_span"])

                # 插入图片
                excel_img = ExcelImage(img_path)
                excel_img.width = cfg["width"]
                excel_img.height = cfg["height"]

                target_cell = f"{get_column_letter(current_col)}{current_row}"
                ws.add_image(excel_img, target_cell)

                # 调用合并单元格和画边框的函数
                self._apply_defect_border(ws, current_row, current_col, cfg["row_span"], cfg["col_span"])

            print(f"✓ 成功在首页插入 {len(unique_defect_images)} 张缺陷图")

        except Exception as e:
            print(f"✗ 插入首页缺陷图片失败: {e}")

    def _apply_defect_border(self, ws, row, col, r_span, c_span):
        """为缺陷图片区域添加边框并合并"""
        from openpyxl.styles import Border, Side
        medium_side = Side(style='medium', color="000000")
        border = Border(top=medium_side, left=medium_side, right=medium_side, bottom=medium_side)

        # 合并区域：例如 B54:E67
        # start_column=2, c_span=4 -> end_column = 2 + 4 - 1 = 5 (E列)
        ws.merge_cells(
            start_row=row,
            start_column=col,
            end_row=row + r_span - 1,
            end_column=col + c_span - 1
        )

        # 遍历合并区域应用边框（openpyxl合并后必须逐个单元格设边框才有完整框线）
        for r in range(row, row + r_span):
            for c in range(col, col + c_span):
                ws.cell(row=r, column=c).border = border

    def create_thumbnail(self, image_path, size=(200, 150)):
        """创建缩略图"""
        try:
            img = Image.open(image_path)
            img.thumbnail(size, Image.Resampling.LANCZOS)
            return img
        except Exception as e:
            print(f"✗ 创建缩略图失败 {image_path}: {e}")
            return None

    def insert_images_to_excel(self, step_images_mapping,po_number):
        """
        将图片插入到Reference pictures工作表
        step_images_mapping格式: {
            'Step 1': ['path/to/image1.jpg', ...],
            'Step 5（1）': ['path/to/image2.jpg', ...],
            ...
        }
        """
        try:
            # 1. 初始化图片工作表
            if 'Reference pictures' not in self.wb.sheetnames:
                ws_pics = self.wb.create_sheet(f"Reference pictures {po_number}")
            else:
                ws_pics = self.wb[f"Reference pictures {po_number}"]

            # 清空原有内容
            ws_pics.delete_rows(1, ws_pics.max_row + 1)

            # 关闭网格线
            ws_pics.sheet_view.showGridLines = False

            # 2. 定义基础样式
            # 字体样式
            base_font = Font(
                name=config.FONT_CONFIG["name"],
                size=config.FONT_CONFIG["size"],
                bold=config.FONT_CONFIG["bold"],
                color=config.FONT_CONFIG["color"]
            )
            # 黑色细边框
            border_side = Side(
                style=config.BORDER_CONFIG["style"],
                color=config.BORDER_CONFIG["color"]
            )
            black_border = Border(
                left=border_side, right=border_side,
                top=border_side, bottom=border_side
            )
            # 对齐方式（垂直居中）
            align = Alignment(vertical="center")

            # 3. 获取基础数据（PO号/SKU/日期/检验员）
            ws_main = self.wb['出货检查表']
            po_number = ws_main['G5'].value or "PO-UNKNOWN"
            sku = ws_main['C6'].value or ""
            inspection_date = ws_main['G4'].value or ""
            inspector = ws_main['C4'].value or ""

            # 4. 填充标题行（B1）
            current_row = 1
            title_text = config.STEP_TEXT["title"].format(po_number=po_number)
            ws_pics[f'B{current_row}'] = title_text
            ws_pics[f'B{current_row}'].font = base_font

            # 5. 填充SKU（B2）
            current_row += 1
            ws_pics[f'B{current_row}'] = f"{config.STEP_TEXT['sku_label']}{sku}"
            ws_pics[f'B{current_row}'].font = base_font

            # 6. 填充日期（F2）和检验员（F3/G3）
            ws_pics[f'F{current_row}'] = config.STEP_TEXT["date_label"]
            ws_pics[f'G{current_row}'] = inspection_date
            ws_pics[f'F{current_row}'].font = base_font
            ws_pics[f'G{current_row}'].font = base_font

            current_row += 1
            ws_pics[f'F{current_row}'] = config.STEP_TEXT["inspector_label"]
            ws_pics[f'G{current_row}'] = inspector
            ws_pics[f'F{current_row}'].font = base_font
            ws_pics[f'G{current_row}'].font = base_font

            # 7. 填充Step1-Step4文本+图片
            step_order = ["Step 1", "Step 2", "Step 3", "Step 4"]
            for step in step_order:
                current_row += 1
                # 统一键名：Step 1 → step1
                step_key = step.lower().replace(" ", "")  # 转为step1/step2...
                step_text = config.STEP_TEXT.get(step_key, f"{step}) 无描述")
                ws_pics[f'B{current_row}'] = step_text
                ws_pics[f'B{current_row}'].font = base_font

                # 插入Step图片（横向排列，带边框）
                images = step_images_mapping.get(step, [])
                if images:
                    current_row += 1
                    self._insert_images_with_border(
                        ws_pics=ws_pics,
                        start_row=current_row,
                        start_col=2,  # B列开始
                        images=images,
                        border=black_border,
                        font=base_font,
                        align=align
                    )
                    # 图片行高适配
                    ws_pics.row_dimensions[current_row].height = config.IMAGE_CONFIG["row_height"]
                    current_row += 1  # 图片后空一行

            # 8. 填充Step5文本+子项+图片
            current_row += 1
            # Step5主标题
            ws_pics[f'B{current_row}'] = config.STEP_TEXT["step5_title"]
            ws_pics[f'B{current_row}'].font = base_font

            # Step5子项文本（a-j）
            step5_sub_items = [
                "step5_1", "step5_2", "step5_3", "step5_4", "step5_5",
                "step5_6", "step5_7", "step5_8", "step5_9", "step5_10", "step5_11"
            ]
            for sub_item in step5_sub_items:
                current_row += 1
                # 写入子项文本
                sub_text = config.STEP_TEXT.get(sub_item, f"{sub_item}: 无描述")
                ws_pics[f'B{current_row}'] = sub_text
                ws_pics[f'B{current_row}'].font = base_font

                # 匹配Step5细分图片
                target_step = config.STEP5_IMAGE_MAP.get(sub_item)
                if target_step:
                    print(f"匹配到Step5子项 {sub_item} → 图片步骤 {target_step}")
                    # 插入Step5细分图片
                    if step_images_mapping.get(target_step):
                        current_row += 1
                        self._insert_images_with_border(
                            ws_pics=ws_pics,
                            start_row=current_row,
                            start_col=2,
                            images=step_images_mapping[target_step],
                            border=black_border,
                            font=base_font,
                            align=align
                        )
                        ws_pics.row_dimensions[current_row].height = config.IMAGE_CONFIG["row_height"]
                        current_row += 1  # 图片后空一行
                    else:
                        print(f"Step5子项 {sub_item} 对应步骤 {target_step} 无图片")
                else:
                    print(f" Step5子项 {sub_item} 无对应的图片步骤映射")
            return True
        except Exception as e:
            print(f"✗ 插入图片到Excel失败: {e}")
            return False

    def _insert_images_with_border(self, ws_pics, start_row, start_col, images, border, font, align):
        if not images:
            return

        fixed_col_gap = config.IMAGE_CONFIG["fixed_col_gap"]
        fixed_col_width = config.IMAGE_CONFIG["fixed_col_width"]

        current_col = start_col
        img_col_list = []
        # 记录所有涉及的列（包括间隔列），确保边框覆盖完整
        all_involved_cols = []

        for img_path in images:
            try:
                img = ExcelImage(img_path)
                img.width = config.IMAGE_CONFIG["width"]
                img.height = config.IMAGE_CONFIG["height"]
                img_cell = f"{get_column_letter(current_col)}{start_row}"
                ws_pics.add_image(img, img_cell)

                # 记录图片所在列+后续间隔列（解决列覆盖不全）
                img_col_list.append(current_col)
                # 标记当前列到下一张图片前的所有列
                next_col = current_col + fixed_col_width // 6 + fixed_col_gap
                all_involved_cols.extend(range(current_col, next_col))

                # 强制设置当前列宽
                ws_pics.column_dimensions[get_column_letter(current_col)].width = fixed_col_width

                # 修正列偏移计算
                current_col = next_col
            except Exception as e:
                print(f"无法插入图片 {img_path}: {e}")
                # 跳过失败图片，列偏移继续（避免后续图片列错位）
                current_col += config.IMAGE_CONFIG["col_offset_step"] + fixed_col_gap
                continue

        if img_col_list:
            # 方案1：用实际涉及的所有列确定合并范围（推荐）
            start_col_border = min(img_col_list)
            # 取最后一张图片的结束列（而非列索引），确保覆盖所有图片
            end_col_border = max(all_involved_cols) if all_involved_cols else img_col_list[-1]

            # 合并范围：覆盖所有图片+间隔列
            merge_range = f"{get_column_letter(start_col_border)}{start_row}:{get_column_letter(end_col_border)}{start_row}"
            ws_pics.merge_cells(merge_range)

            # 方案2：强制设置合并单元格的所有边框（解决样式渲染不全）
            border_cell = f"{get_column_letter(start_col_border)}{start_row}"
            # 1. 基础样式应用
            ws_pics[border_cell].border = border
            ws_pics[border_cell].font = font
            ws_pics[border_cell].alignment = align

            # 2. 遍历合并范围的所有单元格，强制应用边框（关键修复）
            for col in range(start_col_border, end_col_border + 1):
                cell = f"{get_column_letter(col)}{start_row}"
                ws_pics[cell].border = border
                ws_pics[cell].font = font
                ws_pics[cell].alignment = align

            # 3. 强制刷新行高/列宽（避免Excel渲染异常）
            ws_pics.row_dimensions[start_row].height = config.IMAGE_CONFIG["row_height"]
            for col in range(start_col_border, end_col_border + 1):
                ws_pics.column_dimensions[get_column_letter(col)].width = fixed_col_width

    def generate_report_no(self):
        """自动生成报告编号"""
        now = datetime.now()
        return f"OI{now.year % 100:02d}{now.month:02d}{now.day:02d}-{now.hour:02d}{now.minute:02d}"

    def save_report(self, output_path):
        """保存报告"""
        try:
            self.wb.save(output_path)
            print(f"✓ 报告保存成功: {output_path}")
            return True
        except Exception as e:
            print(f"✗ 保存报告失败: {e}")
            return False


class InspectionReportGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("USTC圆饼锁报告生成工具")
        self.root.geometry("1100x700")
        self.root.minsize(800, 600)

        self.generator = InspectionReportGenerator()
        self.selected_images = {}
        self.image_checkbuttons = {}

        # --- 修复点：路径逻辑只保留一份 ---
        if getattr(sys, 'frozen', False):
            # 如果是打包后的 exe，获取 exe 所在的实际文件夹路径
            self.base_dir = os.path.dirname(sys.executable)
        else:
            # 如果是直接运行 .py 脚本
            self.base_dir = os.path.dirname(os.path.abspath(__file__))
        default_template_name = "模板.xlsx"
        self.default_template_path = os.path.join(self.base_dir, default_template_name)

        # --- 修复点：只执行一次初始化逻辑 ---
        self.setup_styles()
        self.create_widgets()

        # 自动尝试加载默认模板
        if os.path.exists(self.default_template_path):
            self.template_var.set(self.default_template_path)
            self.generator.load_template(self.default_template_path)

    def setup_styles(self):
        """设置界面样式"""
        style = ttk.Style()
        style.configure('Title.TLabel', font=('微软雅黑', 16, 'bold'))
        style.configure('Step.TLabel', font=('微软雅黑', 11, 'bold'), foreground='blue')
        style.configure('Accent.TButton', font=('微软雅黑', 10, 'bold'))

    def create_widgets(self):
        """创建支持自适应滚动的主界面组件"""
        # 最外层主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 1. 创建左右分割的水平窗格 (PanedWindow)，允许用户手动调节左右比例且支持自适应
        pw = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        pw.pack(fill=tk.BOTH, expand=True)

        # --- 左侧自适应容器 ---
        left_container = ttk.LabelFrame(pw, text="报告信息", padding="5")
        left_canvas = tk.Canvas(left_container, highlightthickness=0)
        left_scrollbar = ttk.Scrollbar(left_container, orient="vertical", command=left_canvas.yview)
        self.left_scrollable_frame = ttk.Frame(left_canvas)

        self.left_scrollable_frame.bind(
            "<Configure>",
            lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all"))
        )
        left_canvas.create_window((0, 0), window=self.left_scrollable_frame, anchor="nw")
        left_canvas.configure(yscrollcommand=left_scrollbar.set)

        left_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        left_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        pw.add(left_container, weight=1)

        # --- 右侧自适应容器 ---
        right_container = ttk.LabelFrame(pw, text="图片管理", padding="5")
        right_canvas = tk.Canvas(right_container, highlightthickness=0)
        right_scrollbar = ttk.Scrollbar(right_container, orient="vertical", command=right_canvas.yview)
        self.right_scrollable_frame = ttk.Frame(right_canvas)

        self.right_scrollable_frame.bind(
            "<Configure>",
            lambda e: right_canvas.configure(scrollregion=right_canvas.bbox("all"))
        )
        right_canvas.create_window((0, 0), window=self.right_scrollable_frame, anchor="nw")
        right_canvas.configure(yscrollcommand=right_scrollbar.set)

        right_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        right_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        pw.add(right_container, weight=1)

        # 在滚动的内部框架中创建具体内容
        self.create_data_input(self.left_scrollable_frame)
        self.create_image_manager(self.right_scrollable_frame)

        # 底部固定按钮区
        self.create_bottom_buttons(main_frame)

    def on_sku_selected(self, event=None):
        sku = self.sku_var.get()

        # 每次切换 SKU，先清空旧状态
        self.drawing_var.set("")

        for model, drawing in config.DRAWING_RULES.items():
            if model in sku:
                self.drawing_var.set(drawing)
                return

        # 明确兜底（防呆）
        self.drawing_var.set("【未匹配图纸】")

    def create_data_input(self, parent):
        """创建数据输入内容（此部分保持原有逻辑，仅 parent 指向滚动容器）"""
        # 模板选择
        ttk.Label(parent, text="Excel模板:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.template_var = tk.StringVar(value=getattr(self, 'default_template_path', ""))
        ttk.Entry(parent, textvariable=self.template_var, width=30).grid(
            row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=(5, 0))
        ttk.Button(parent, text="浏览...", command=self.browse_template).grid(
            row=0, column=2, pady=5, padx=(5, 0))

        # 检验员
        ttk.Label(parent, text="检验员:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.inspector_var = tk.StringVar()
        ttk.Entry(parent, textvariable=self.inspector_var, width=30).grid(
            row=1, column=1, columnspan=2, sticky=tk.W, pady=5, padx=(5, 0))

        # 检验日期
        ttk.Label(parent, text="检验日期:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y/%m/%d"))
        ttk.Entry(parent, textvariable=self.date_var, width=30).grid(
            row=2, column=1, columnspan=2, sticky=tk.W, pady=5, padx=(5, 0))

        # 订单号
        ttk.Label(parent, text="客户订单号:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.po_var = tk.StringVar()
        ttk.Entry(parent, textvariable=self.po_var, width=30).grid(
            row=3, column=1, columnspan=2, sticky=tk.W, pady=5, padx=(5, 0))

        # 料号
        ttk.Label(parent, text="料号:").grid(row=4, column=0, sticky=tk.W, pady=5)

        self.sku_var = tk.StringVar(value=config.SKU_OPTIONS[0])

        self.sku_combo = ttk.Combobox(
            parent,
            textvariable=self.sku_var,
            values=config.SKU_OPTIONS,
            width=28,
            state="readonly"
        )
        self.sku_combo.grid(row=4, column=1, columnspan=2, sticky=tk.W, pady=5, padx=(5, 0))
        self.sku_combo.bind("<<ComboboxSelected>>", self.on_sku_selected)

        # 客户信息
        ttk.Label(parent, text="客户:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.customer_var = tk.StringVar(value="Master Lock")
        ttk.Entry(parent, textvariable=self.customer_var, width=30).grid(
            row=5, column=1, columnspan=2, sticky=tk.W, pady=5, padx=(5, 0))

        # 客户图纸及版本号
        ttk.Label(parent, text="客户图纸及版本号:").grid(row=6, column=0, sticky=tk.W, pady=5)
        self.drawing_var = tk.StringVar()
        ttk.Entry(
            parent,
            textvariable=self.drawing_var,
            width=30,
            state="readonly"
        ).grid(
            row=6, column=1, columnspan=2, sticky=tk.W, pady=5, padx=(5, 0))

        # 批准人
        ttk.Label(parent, text="批准人:").grid(row=7, column=0, sticky=tk.W, pady=5)
        self.approver_var = tk.StringVar(value="Gary Tu")
        ttk.Entry(parent, textvariable=self.approver_var, width=30).grid(
            row=7, column=1, columnspan=2, sticky=tk.W, pady=5, padx=(5, 0))

        # 批准日期
        ttk.Label(parent, text="批准日期:").grid(row=8, column=0, sticky=tk.W, pady=5)
        self.approval_date_var = tk.StringVar(value=datetime.now().strftime("%Y/%m/%d"))
        ttk.Entry(parent, textvariable=self.approval_date_var, width=30).grid(
            row=8, column=1, columnspan=2, sticky=tk.W, pady=5, padx=(5, 0))

        # 出货数量
        ttk.Label(parent, text="出货数量:").grid(row=9, column=0, sticky=tk.W, pady=5)
        self.quantity_var = tk.StringVar()
        ttk.Entry(parent, textvariable=self.quantity_var, width=30).grid(
            row=9, column=1, columnspan=2, sticky=tk.W, pady=5, padx=(5, 0))

        # 计划出货日期
        ttk.Label(parent, text="计划出货日期:").grid(row=10, column=0, sticky=tk.W, pady=5)
        self.ship_date_var = tk.StringVar(value=datetime.now().strftime("%Y/%m/%d"))
        ttk.Entry(parent, textvariable=self.ship_date_var, width=30).grid(
            row=10, column=1, columnspan=2, sticky=tk.W, pady=5, padx=(5, 0))

        # 报告编号
        ttk.Label(parent, text="报告编号:").grid(row=11, column=0, sticky=tk.W, pady=5)
        self.report_no_var = tk.StringVar()
        ttk.Entry(parent, textvariable=self.report_no_var, width=20).grid(
            row=11, column=1, sticky=tk.W, pady=5, padx=(5, 0))
        ttk.Button(parent, text="自动生成", command=self.generate_report_no, width=10).grid(
            row=11, column=2, pady=5, padx=(5, 0))

        # 缺陷记录区域标题
        ttk.Label(parent, text="缺陷记录 (选填):", font=('微软雅黑', 10, 'bold')).grid(
            row=12, column=0, columnspan=3, sticky=tk.W, pady=(20, 5))

        # 缺陷记录表头及行逻辑保持不变
        headers = ["序号", "缺陷描述", "致命", "严重", "轻微"]
        for col, header in enumerate(headers):
            ttk.Label(parent, text=header, font=('微软雅黑', 9, 'bold')).grid(
                row=13, column=col, padx=2, pady=2)

        self.defect_vars = []
        for i in range(8):
            row_vars = []
            row_num = 14 + i
            ttk.Label(parent, text=str(i + 1)).grid(row=row_num, column=0, padx=2, pady=2)
            desc_var = tk.StringVar()
            ttk.Entry(parent, textvariable=desc_var, width=20).grid(row=row_num, column=1, padx=2, pady=2, sticky=tk.W)
            row_vars.append(desc_var)
            for col in range(2, 5):
                var = tk.StringVar(value="0")
                ttk.Entry(parent, textvariable=var, width=5).grid(row=row_num, column=col, padx=2, pady=2)
                row_vars.append(var)
            self.defect_vars.append(row_vars)
        self.on_sku_selected()

    def create_image_manager(self, parent):
        """创建图片管理内容（指向 parent）"""
        folder_frame = ttk.Frame(parent)
        folder_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(folder_frame, text="图片文件夹:").pack(side=tk.LEFT)
        self.image_folder_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.image_folder_var, width=30).pack(side=tk.LEFT, padx=5)
        ttk.Button(folder_frame, text="浏览...", command=self.browse_image_folder).pack(side=tk.LEFT)
        ttk.Button(folder_frame, text="扫描", command=self.scan_images).pack(side=tk.LEFT, padx=5)

        # 内部图片预览区域（嵌套 Canvas 保持原有逻辑）
        preview_container = ttk.Frame(parent)
        preview_container.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(preview_container, bg='#f0f0f0', height=400)
        img_scrollbar = ttk.Scrollbar(preview_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=img_scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        img_scrollbar.pack(side="right", fill="y")

        # 统计信息
        self.create_stats_area(parent)

    def create_stats_area(self, parent):
        """原有的步骤统计显示逻辑"""
        stats_frame = ttk.Frame(parent)
        stats_frame.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(stats_frame, text="各步骤图片统计:", font=('微软雅黑', 10, 'bold')).pack(anchor=tk.W)

        self.step_counts = {}
        steps = ['Step 1', 'Step 2', 'Step 3', 'Step 4', 'Step 5（1）', 'Step 5（2）', 'Step 5（3）', 'Step 5（4）',
                 'Step 5（5）',]

        # 使用流式布局防止统计块在横向消失
        flow_frame = ttk.Frame(stats_frame)
        flow_frame.pack(fill=tk.X)
        for i, step in enumerate(steps):
            f = ttk.Frame(flow_frame, relief=tk.RIDGE, padding="2")
            f.grid(row=i // 3, column=i % 3, padx=2, pady=2, sticky=tk.NSEW)
            ttk.Label(f, text=step, font=('微软雅黑', 8)).pack()
            cv = tk.StringVar(value="0张")
            ttk.Label(f, textvariable=cv, font=('微软雅黑', 8, 'bold')).pack()
            self.step_counts[step] = cv

    def create_bottom_buttons(self, parent):
        """底部按钮区"""
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(button_frame, text="生成报告", command=self.generate_report, style='Accent.TButton').pack(
            side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="清除数据", command=self.clear_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="退出", command=self.root.quit).pack(side=tk.RIGHT, padx=5)

    def browse_template(self):
        """浏览选择Excel模板"""
        filename = filedialog.askopenfilename(
            title="选择Excel模板",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if filename:
            self.template_var.set(filename)
            if self.generator.load_template(filename):
                messagebox.showinfo("成功", f"模板加载成功: {Path(filename).name}")

    def browse_image_folder(self):
        """浏览选择图片文件夹"""
        folder = filedialog.askdirectory(title="选择图片文件夹")
        if folder:
            self.image_folder_var.set(folder)

    def scan_images(self):
        """扫描图片文件夹（最优自适应布局：确保右侧操作项始终可见）"""
        folder = self.image_folder_var.get()
        if not folder:
            messagebox.showwarning("警告", "请先选择图片文件夹")
            return

        # 扫描图片
        images = self.generator.scan_images_folder(folder)
        if not images:
            messagebox.showinfo("提示", "未找到图片文件")
            return

        # 清空预览区
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        # 清空之前的选择
        self.image_checkbuttons.clear()

        # 初始化步骤计数
        step_image_counts = {k: 0 for k in [
            'Step 1', 'Step 2', 'Step 3', 'Step 4',
            'Step 5（1）', 'Step 5（2）', 'Step 5（3）', 'Step 5（4）', 'Step 5（5）'
        ]}

        # 配置滚动容器的列权重，使其横向铺满
        self.scrollable_frame.columnconfigure(0, weight=1)

        row = 0
        for img_data in images:
            # 1. 创建条目主框架：必须 sticky=EW 以铺满宽度
            frame = ttk.Frame(self.scrollable_frame, relief=tk.RIDGE, padding="5")
            frame.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=3, padx=5)

            # 关键：配置框架内部列权重
            # Column 0: 缩略图 (固定)
            # Column 1: 信息区 (自动拉伸)
            # Column 2: 操作区 (固定在右侧)
            frame.columnconfigure(1, weight=1)

            # 2. 缩略图
            thumbnail = self.generator.create_thumbnail(img_data['path'], size=(120, 90))
            if thumbnail:
                photo = ImageTk.PhotoImage(thumbnail)
                img_label = ttk.Label(frame, image=photo)
                img_label.image = photo
                img_label.grid(row=0, column=0, rowspan=2, padx=(0, 10), sticky=tk.NW)

                # 3. 中间信息区域
                info_frame = ttk.Frame(frame)
                info_frame.grid(row=0, column=1, sticky=(tk.W, tk.E))

                # 限制 wraplength 以触发自动换行，防止无限撑开
                file_name_label = ttk.Label(info_frame,
                                            text=f"文件名: {img_data['filename']}",
                                            font=('微软雅黑', 9, 'bold'),
                                            wraplength=180)  # 这里的像素值会随框架缩放起作用
                file_name_label.pack(anchor=tk.W, fill=tk.X)

                ttk.Label(info_frame, text=f"原始分类: {img_data['step']}",
                          font=('微软雅黑', 9)).pack(anchor=tk.W)

                # 4. 步骤分配
                step_frame = ttk.Frame(frame)
                step_frame.grid(row=1, column=1, sticky=tk.W, pady=(5, 0))
                ttk.Label(step_frame, text="重新分配到:", font=('微软雅黑', 9)).pack(side=tk.LEFT)

                step_options = [
                    'Step 1', 'Step 2', 'Step 3', 'Step 4',
                    'Step 5（1）', 'Step 5（2）', 'Step 5（3）', 'Step 5（4）', 'Step 5（5）'
                ]
                step_var = tk.StringVar(value=img_data['step'])
                step_combo = ttk.Combobox(step_frame, textvariable=step_var, values=step_options, width=12,
                                          state='readonly')
                step_combo.pack(side=tk.LEFT, padx=(5, 10))

                # 5. 右侧操作区域：使用 sticky=E 强制靠右
                action_frame = ttk.Frame(frame)
                action_frame.grid(row=0, column=2, rowspan=2, padx=(10, 5), sticky=tk.E)

                check_var = tk.BooleanVar(value=True)
                check_btn = ttk.Checkbutton(action_frame, text="使用", variable=check_var)
                check_btn.pack(anchor=tk.E)  # 靠右对齐

                defect_var = tk.BooleanVar(value=False)
                if any(word.lower() in img_data['filename'].lower() for word in config.DEFECT_WORDS):
                    defect_var.set(True)

                defect_btn = ttk.Checkbutton(action_frame, text="设为缺陷图", variable=defect_var)
                defect_btn.pack(anchor=tk.E, pady=(5, 0))

                # 保存数据引用
                self.image_checkbuttons[img_data['path']] = {
                    'checkbox': check_var,
                    'step': step_var,
                    'defect_var': defect_var,
                    'data': img_data
                }

                if img_data['step'] in step_image_counts:
                    step_image_counts[img_data['step']] += 1

            row += 1

        # 更新统计显示
        for step, count_var in self.step_counts.items():
            count_var.set(f"{step_image_counts.get(step, 0)}张")

    def generate_report_no(self):
        """生成报告编号"""
        report_no = self.generator.generate_report_no()
        self.report_no_var.set(report_no)
        messagebox.showinfo("报告编号", f"已生成报告编号: {report_no}")

    def get_defects_data(self):
        """获取缺陷数据"""
        defects = []
        for row_vars in self.defect_vars:
            description = row_vars[0].get().strip()
            if description:  # 只添加有描述的缺陷
                try:
                    critical = int(row_vars[1].get() or 0)
                    major = int(row_vars[2].get() or 0)
                    minor = int(row_vars[3].get() or 0)

                    defects.append({
                        'description': description,
                        'critical': critical,
                        'major': major,
                        'minor': minor
                    })
                except ValueError:
                    continue
        return defects

    def get_selected_images(self):
        """获取选择的图片并按步骤分组（适配Step5细分）"""
        step_images = {
            'Step 1': [], 'Step 2': [], 'Step 3': [], 'Step 4': [],
            'Step 5（1）': [], 'Step 5（2）': [], 'Step 5（3）': [], 'Step 5（4）': [], 'Step 5（5）': []
        }

        for img_info in self.image_checkbuttons.values():
            if img_info['checkbox'].get():
                step = img_info['step'].get()
                if step in step_images:
                    step_images[step].append(img_info['data']['path'])

        return step_images

    def generate_report(self):
        """生成报告主逻辑（含缺陷图重命名与首页自动插入）"""
        # 1. 基础验证
        template_path = self.template_var.get()
        if not template_path:
            messagebox.showerror("错误", "请选择Excel模板")
            return

        # 每次生成报告都重新加载模板（重置 Workbook）
        if not self.generator.load_template(template_path):
            messagebox.showerror("错误", "加载模板失败")
            return
        # 校验出货数量
        try:
            ship_quantity = int(self.quantity_var.get() or 0)
            if ship_quantity <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("错误", "请输入有效的出货数量")
            return

        # --- 核心新增：处理缺陷图物理重命名并收集路径 ---
        collected_defect_paths = []
        for original_path, info in self.image_checkbuttons.items():
            if info['checkbox'].get():  # 如果勾选了“使用”
                current_path = original_path

                # 如果勾选了“设为缺陷图”
                if info['defect_var'].get():
                    p = Path(original_path)
                    # 执行物理重命名逻辑
                    if not p.name.startswith("(缺陷)"):
                        new_name = f"(缺陷){p.name}"
                        new_path = p.with_name(new_name)
                        try:
                            os.rename(original_path, new_path)
                            current_path = str(new_path)  # 更新路径为重命名后的
                            info['data']['path'] = current_path  # 同步更新内存数据
                        except Exception as e:
                            print(f"重命名失败: {e}")

                    # 将该路径加入缺陷列表（无论是新命名的还是原本就带缺陷字样的）
                    collected_defect_paths.append(current_path)

        # 2. 填充基本信息和抽样计划
        data = {
            'inspector': self.inspector_var.get(),
            'inspection_date': self.date_var.get(),
            'po_number': self.po_var.get(),
            'sku': self.sku_var.get(),
            'ship_date': self.ship_date_var.get(),
            'ship_quantity': ship_quantity,
            'report_no': self.report_no_var.get() or self.generator.generate_report_no(),
            'customer': self.customer_var.get(),
            'drawing_no': self.drawing_var.get(),
            'approver': self.approver_var.get(),
            'approval_date': self.approval_date_var.get()
        }
        self.generator.fill_basic_info(data)

        # 3. 填充文字缺陷记录
        defects = self.get_defects_data()
        if defects:
            self.generator.add_defect_records(defects)

        # --- 核心新增：将收集到的缺陷图插入到 Excel 首页 ---
        if collected_defect_paths:
            self.generator.defect_images = collected_defect_paths
            self.generator._insert_defect_images()

        # 4. 插入常规图片页（Step 1-5）
        step_images = self.get_selected_images()
        if any(step_images.values()):
            po_number = self.po_var.get().strip() or "PO"
            self.generator.insert_images_to_excel(step_images, po_number)
        # --- 根据 SKU 判断型号，用于文件名 ---
        sku = self.sku_var.get()
        po_number = self.po_var.get().strip() or "PO"

        if "M40" in sku:
            model_prefix = "M40"
        elif "M50" in sku:
            model_prefix = "M50"
        else:
            model_prefix = "MODEL"
        # 5. 保存文件
        default_name = f"{model_prefix}_{po_number}.xlsx"
        output_file = filedialog.asksaveasfilename(
            title="保存报告",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel文件", "*.xlsx")]
        )

        if output_file:
            if self.generator.save_report(output_file):
                messagebox.showinfo("成功", f"报告已生成！\n缺陷图已重命名并同步至首页。")
                if messagebox.askyesno("打开", "是否打开生成的报告？"):
                    os.startfile(output_file)

    def clear_data(self):
        """清除所有数据"""
        # 清除基本信息
        self.inspector_var.set("")
        self.po_var.set("")
        self.quantity_var.set("")
        self.report_no_var.set("")
        self.customer_var.set("Master Lock")  # 重置为默认值
        self.drawing_var.set("64678 Rev.J")  # 重置为默认值
        self.approver_var.set("Gary Tu")  # 重置为默认值
        self.approval_date_var.set(datetime.now().strftime("%Y/%m/%d"))  # 重置为当前日期
        # 清除缺陷记录
        for row_vars in self.defect_vars:
            for var in row_vars:
                if var == row_vars[0]:  # 描述字段
                    var.set("")
                else:
                    var.set("0")

        # 清除图片预览
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        # 重置步骤计数（包含Step5细分）
        for count_var in self.step_counts.values():
            count_var.set("0张")

        # 清空图片选择
        self.image_checkbuttons.clear()

        messagebox.showinfo("清除", "所有数据已清除")


def main():
    """主函数"""
    # 检查依赖
    try:
        import openpyxl
        from PIL import Image
    except ImportError as e:
        print(f"缺少依赖库: {e}")
        print("请安装所需库: pip install openpyxl Pillow")
        messagebox.showerror("错误", f"缺少依赖库: {e}\n请运行: pip install openpyxl Pillow")
        return

    # 创建GUI
    root = tk.Tk()
    app = InspectionReportGUI(root)

    # 启动主循环
    root.mainloop()


if __name__ == "__main__":
    main()